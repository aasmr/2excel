import os, openpyxl, re, sys, urllib.request, urllib.error, subprocess
from openpyxl.styles import Alignment, Font
from PyPDF2 import PdfFileReader
from _ast import If
print('© Смирнов Алексей, 2018')
def update():
    try:
        version_file=urllib.request.urlopen('http://madhound.ru/downloads/2excel/version.txt').readline()
        ver=version_file.decode('utf-8')
        ver=ver.rstrip().split('=')
        ver=ver[1]
        if ver>version:
            #bat=open('remove.bat', 'w')
            #bat.write('@echo off\nping -n 3 localhost > nul\ndel _2excel.exe /F\ndel %0')
            #bat.close()
            os.rename('2excel.exe', '_2excel.exe')
            new_ver=urllib.request.urlopen('http://madhound.ru/downloads/2excel/2excel.exe').read()
            file=open('2excel.exe', 'wb')
            file.write(new_ver)
            file.close()
            #print("Программа обновилась успешно")
            subprocess.Popen('echo 123')
            sys.exit()
            #input('Программа обновлена, сейчас она перезапустится')
        else:
            input('У вас установлена самая свежая версия программы...')
    except Exception as e:
        print(e)
##        if e==urllib.error.URLError:
##            print(e)
##            print(e.reason)
##        else:
##            print('Error', e.code)
##            print(e.reason)

def post_update():
    subprocess.Call('nping -n 3 localhost > nul')
    os.remove('_2excel.exe')
    print("Программа обновилась успешно")
    sys.exit()
def format_A(w, h):
    w_ls=[210, 297, 420, 594, 841, 1189]
    h_ls=[210, 297, 420, 594, 841, 1189]

    flag=0

    for elem in h_ls:
        if h==elem:
            if h>w:
                flag=1
                if [w, h] > [836, 1184] and [w, h] < [846, 1194]:
                    form='A0'
                elif [w, h] > [589, 836] and [w, h] < [599, 846]:
                    form='A1'
                elif [w, h] > [415, 589] and [w, h] < [425, 599]:
                    form='A2'
                elif [w, h] > [292, 415] and [w, h] < [302, 425]:
                    form='A3'
                elif [w, h] > [205, 292] and [w, h] < [215, 302]:
                    form='A4'
                else:
                    form='Ax'
                    
            if h<w and w==w_ls[h_ls.index(h)+1]:
                flag=1
                if [h, w] > [836, 1184] and [h, w] < [846, 1194]:
                    form='A0'
                elif [h, w] > [589, 836] and [h, w] < [599, 846]:
                    form='A1'
                elif [h, w] > [415, 589] and [h, w] < [425, 599]:
                    form='A2'
                elif [h, w] > [292, 415] and [h, w] < [302, 425]:
                    form='A3'
                elif [h, w] > [205, 292] and [h, w] < [215, 302]:
                    form='A4'
                else:
                    form='Ax'
                    
            if h<w and w!=w_ls[h_ls.index(h)+1]:
                flag=1
                if h == 1189:
                    form='A0x'
                elif h == 841:
                    form='A1x'
                elif h == 594:
                    form='A2x'
                elif h == 420:
                    form='A3x'
                elif h == 297:
                    form='A4x'
                else:
                    form='Ax'
                x=str(int(w/w_ls[h_ls.index(h)-1]))
                form=form+x
                
        if flag == 0:
            form='Ax'
    return form
def search(string):
    end=0
    flag=0
    ls=[]

    spec_des_1=r'\d+[-]{1}\d+([-]*\d+)*( )*(СБ|СП|SB|SP)*'
    result_des=re.match(spec_des_1, string)

    if result_des:
        end=result_des.end()
        ls.append(result_des.group())
    else:
        spec_des_1=r'\d+[.]{1}\d+[-]{1}\d+[.]{1}\d+( )*(СБ|СП|SB|SP)*'
        result_des=re.match(spec_des_1, string)
        if result_des:
            end=result_des.end()
            ls.append(result_des.group())
        else:
            spec_des_1=r'\w+[_]{1}\d+[_]{1}\d+[_]{1}\d+(\s|_)*(СБ|СП|SB|SP)*'
            result_des=re.match(spec_des_1, string)
            if result_des:
                end=result_des.end()
                ls.append(result_des.group())
            else:
                spec_des_1=r'\w+[\s]{1}\d+[.]{1}\d+[.]{1}\d+(\s|_)*(СБ|СП|SB|SP)*'
                result_des=re.match(spec_des_1, string)
                if result_des:
                    end=result_des.end()
                    ls.append(result_des.group())
                else:
                    ls.append('')

    result_name=string[end:]
    while result_name[0]==' ':
        result_name=result_name[1:]
    if result_name.find('.')!=-1:
        while result_name[-1]!='.':
            result_name=result_name[:-1]
        if result_name[-1]=='.':
            result_name=result_name[:-1]
    ls.append(result_name)

    return ls
    
def parse():
    dims={}
    wb = openpyxl.Workbook()
    ws = wb.active
    align=Alignment(horizontal='center', vertical='center', wrap_text=True, shrink_to_fit=True)
    font=Font('Calibri', 12, bold=True)
    cell=ws.cell(1, 1, 'Обозначение')
    cell.alignment=align
    cell=ws.cell(1, 2, 'Наименование')
    cell.alignment=align
    cell=ws.cell(1, 3, 'Формат')
    cell.alignment=Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=True)
    cell=ws.cell(1, 4, 'Количество листов')
    cell.alignment=align
    dims['A']=max((dims.get('A', 0), len('Обозначение')))
    dims['B']=max((dims.get('B', 0), len('Наименование')))
    dims['C']=max((dims.get('C', 0), len('Формат')))
    dims['D']=max((dims.get('D', 0), len('Количество листов')))

    folder=[]
    for i in os.walk('.\\'):
        folder.append(i)
    k=2
    for adress, dir, file in folder:
        filelist=os.listdir(adress)
        l=1
        flag=0
        for i in filelist:
                        ls=search(i)
                        if re.findall(r'.pdf', i):
                            flag=1
        if flag and adress != '.\\':
                    ws.merge_cells(start_row=k, start_column=1, end_row=k, end_column=4)
                    cell=ws.cell(k, 1, adress)
                    cell.alignment=align
                    cell.font=font
                    k=k+1
                    l=l+1
        for i in filelist:
            ls=search(i)
            if re.findall(r'.pdf', i):
                flag=1
                pdf=PdfFileReader(open(adress+'/'+i, 'rb'), strict=False)
                num=pdf.getNumPages()
                form_ls=[]
                for j in range(num):
                    w=round(float(pdf.getPage(j).mediaBox.upperRight[0])*25.4/72)
                    h=round(float(pdf.getPage(j).mediaBox.upperRight[1])*25.4/72)
                    form=format_A(w, h)
                    if form_ls.count(form)>=1:
                        continue
                    else:
                        form_ls.append(form)
                    if form=='Ax':
                        print('Неопределен формат у файла: ', i,'\nРазмеры в файле:', w, 'x', h)

            if ls[0]!='':
                des=ls[0]
                #print(ls)
                name=ls[1]
                dims['A']=max((dims.get('A', 0), len(des)))
                dims['B']=max((dims.get('B', 0), len(name)))
                cell=ws.cell(k, 1, des)
                #cell.alignment=align
                cell=ws.cell(k, 2, name)
                cell.alignment=Alignment(vertical='center', wrap_text=True, shrink_to_fit=True)
                string=''
                for i in form_ls:
                    string=string+i+', '
                string=string[:-2]
                dims['C']=max((dims.get('C', 0), len(string)))
                dims['D']=max((dims.get('D', 0), len(str(num))))
                cell=ws.cell(k, 3, string)
                cell.alignment=align
                cell=ws.cell(k, 4, num)
                cell.alignment=align
                k=k+1
    wb.save('Опись.xlsx')
    wb=openpyxl.load_workbook('./Опись.xlsx', data_only=True)
    sheet=wb.active
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 0.71
    wb.save('Опись.xlsx')
    return k-l

if __name__ == '__main__':
    version='2.1'
    help_text=''
    changes='\
ver 2.1\n\n\
-Добавлены новые форматы имени файлов.\n\
-Убрана возможность обновления\n\n\
ver 2.0\n\n\
-Добавлена возможность поиска PDF файлов в подкатологах.\n\
-Изменен модуль обработки файлов Excel.\n\
Теперь создается файл .xlsx, в самой таблице теперь столбцы расширены по содержимому.\n\n\
ver 1.2.1\n\n\
-Исправлен вылет программы при обработки некоторых PDF файлов.\n\n\
ver 1.2:\n\n\
-Добавлена возможность обновления программы.\n\
Для того чтобы обновить программу запустите её с аргументом "--update"\n\n\
-Названия форматов теперь не дублируются\n\n\
ver 1.1.1:\n\n\
-Исправлен баг с записью расширений в наименование\n\n\
ver 1.1.0:\n\n\
-Доработан обработчик обозначений и наименований.\n\n\
Теперь обрабатывается обозначение с любым количеством цифр. \
В обозначения так же теперь добавляются литеры "СП", "SB", "SP". \
В наименованиях теперь обрабатываются любые наименования любых форматов файлов.\n\n\
-Добавлена возможность запуска с аргументами --version и --changes.\
В разработке -h, --help, --update.\n\n-Добавлено распознавание составных форматов. \
Теперь проверяется формат каждой страницы в документе.'
    ls=sys.argv
    if len(ls)==1:
        k=parse()-1
        print('Обработано ', k, ' файл(-ов).')
        input('Нажмите любую клавишу для выхода из программы...')
    else:
        for i in ls:
            if i == '--version':
                print(version)
            elif i == '-h' or i == '--help':
                print(help_text)
            elif i == '--changes':
                print(changes)
            #elif i == '--update':
                #update()
            elif i == '--post-update':
                post_update()